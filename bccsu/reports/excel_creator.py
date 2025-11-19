from pathlib import Path
import openpyxl
from openpyxl.reader.excel import ExcelReader
from openpyxl.styles import Alignment
import win32com.client
import tempfile
import json
import logging
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import pandas as pd
from openpyxl.styles import numbers
from datetime import datetime
from openpyxl import Workbook
from time import sleep
import os

Path('logs').mkdir(exist_ok=True)
if not logging.getLogger().handlers:
    log_format = "%(asctime)s | %(levelname)s | %(name)s | %(message)s"
    log_file = Path("logs") / "excel_creator.log"
    file_handler = logging.FileHandler(log_file, encoding="utf-8")
    file_handler.setFormatter(logging.Formatter(log_format))
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(logging.Formatter(log_format))
    logging.basicConfig(level=logging.INFO, handlers=[file_handler, console_handler])

def load_workbook(filename, template=None, *args, **kwargs):
    if template is None:
        template = Path(__file__).parent / 'templates/template.xlsx'
    else:
        template = Path(template)
    destination = Path(filename)
    destination.write_bytes(template.read_bytes())
    return openpyxl.load_workbook(filename, *args, **kwargs)


# Deprecated
def set_page_title(ws, title, pos=None):
    logging.info(f"Writing Title to Excel:\n{title}")
    if pos is None:
        pos = [0, 0]
    cell = ws.cell(row=pos[0] + 1, column=pos[1] + 1, value=title)
    cell.style = 'SAS page title'


def add_note(ws, note, position=None, merge=None):
    logging.info(f"Writing Note to Excel:\n{note}")
    if position is None:
        position = [1, 1]
    position[0] += 1
    position[1] += 1
    cell = ws.cell(row=position[0], column=position[1], value=note)
    if merge:
        ws.merge_cells(start_row=position[0],
                       start_column=position[1],
                       end_row=position[0] + merge[0],
                       end_column=position[1] + merge[1])
    cell.style = 'SAS note'

def add_title(ws, title, position=None):
    logging.info(f"Writing Title to Excel:\n{title}")
    if position is None:
        position = [1, 1]
    position[0] += 1
    position[1] += 1
    cell = ws.cell(row=position[0], column=position[1], value=title)
    cell.style = 'SAS page title'

def get_contiguous_sequences(codes):
    previous_index = [0 for _ in codes[0]]
    final_position = len(codes[0]) - 1
    groups = []
    for col in codes:
        previous_code = col[0]
        start = 0
        sub_groups = []
        for pos, code in enumerate(col):
            if pos != 0:
                if previous_code != code or previous_index[pos] != previous_index[pos - 1]:
                    if pos - start > 1:
                        sub_groups.append([start, pos - 1])
                    start = pos
            previous_code = code
        if final_position - start > 0:
            sub_groups.append([start, final_position])
        groups.append(sub_groups)
        previous_index = col
    return groups


def get_index_codes(index):
    if index.nlevels > 1:
        return index.codes
    else:
        unique_values = index.unique().to_list()
        return [[unique_values.index(value) for value in index]]


def write_table(ws, df, table_start_pos=None, write_index=True, write_headers=True, title=None, percent=False):
    logging.info(f"Writing DataFrame to Excel:\n{df.to_string()}")
    df = df.copy()
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = pd.MultiIndex.from_tuples(
            tuple(str(item) for item in level) for level in df.columns
        )
    else:
        df.columns = df.columns.astype(str)

    if table_start_pos is None:
        table_start_pos = [0, 0]

    table_original_start = table_start_pos

    if write_index:
        index_levels = df.index.nlevels
        index_groupings = get_contiguous_sequences(get_index_codes(df.index))
    else:
        index_levels = 0
        index_groupings = []

    if write_headers:
        column_levels = df.columns.nlevels
        header_groupings = get_contiguous_sequences(get_index_codes(df.columns))
    else:
        column_levels = 0
        header_groupings = []

    if title:
        cell = ws.cell(row=table_start_pos[0] + 1, column=table_start_pos[1] + 1, value=title)
        cell.style = 'SAS Title'
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(start_row=table_start_pos[0] + 1,
                       start_column=table_start_pos[1] + 1,
                       end_row=table_start_pos[0] + 1,
                       end_column=table_start_pos[1] + index_levels + df.shape[1])
        table_start_pos[0] += 1

    if write_index:
        try:
            df_temp = df.reset_index()
        except ValueError:
            df_temp = df.copy()
            df_temp.index.name = ''
            df_temp.reset_index(inplace=True)
    else:
        df_temp = df

    for column_n, (headers, column) in enumerate(df_temp.to_dict().items()):
        if isinstance(headers, str):
            headers = [headers]
        is_index = column_n < index_levels

        current_column = column_n + table_start_pos[1] + 1
        for row_n, value in enumerate(column.values()):
            if row_n == 0 and write_headers:
                for header_pos, header in enumerate(headers):
                    if is_index and (headers[0][:5] in ['level', 'index']):
                        header = ''
                    cell = ws.cell(row=table_start_pos[0] + 1 + header_pos, column=current_column, value=header)
                    cell.style = 'SAS Title'
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            current_row = row_n + table_start_pos[0] + 1 + column_levels
            cell = ws.cell(row=current_row, column=current_column, value=value)
            if is_index:
                cell.style = 'SAS Title'
            else:
                cell.style = 'SAS Data'
                if percent:
                    cell.number_format = numbers.FORMAT_PERCENTAGE_00

    for i, index_grouping in enumerate(index_groupings):
        for group in index_grouping:
            ws.merge_cells(start_row=table_start_pos[0] + 1 + group[0] + column_levels,
                           start_column=table_start_pos[1] + 1 + i,
                           end_row=table_start_pos[0] + 1 + group[1] + column_levels,
                           end_column=table_start_pos[1] + 1 + i)

    for i, index_grouping in enumerate(header_groupings):
        for group in index_grouping:
            ws.merge_cells(start_row=table_start_pos[0] + 1,
                           start_column=table_start_pos[1] + 1 + i + group[0] + index_levels,
                           end_row=table_start_pos[0] + 1,
                           end_column=table_start_pos[1] + 1 + i + group[1] + index_levels)

    if column_levels > 1 or index_levels > 1:
        # Untested. But should work.
        ws.merge_cells(start_row=table_start_pos[0] + 1,
                       start_column=table_start_pos[1] + 1,
                       end_row=table_start_pos[0] + column_levels,
                       end_column=table_start_pos[1] + index_levels)

    # Makes cells background color change

    # for column in ws.columns:
    #     max_length = 0
    #     column = [cell for cell in column if cell.value]  # Checking if the cell has a value
    #     for cell in column:
    #         try:  # Necessary to avoid error on empty cells
    #             if len(str(cell.value)) > max_length:
    #                 max_length = len(cell.value)
    #         except:
    #             pass
    #     adjusted_width = (max_length + 2)  # Additional +2 for a bit of margin
    #     try:
    #         ws.column_dimensions[column[0].column_letter].width = min(adjusted_width, 100)
    #     except:
    #         pass

    return [table_original_start,
            [table_start_pos[0] + column_levels + (1 if title else 0) + df.shape[0],
             table_start_pos[1] + index_levels + df.shape[1]]]


def insert_image(ws, image_path, pos):
    logging.info(f"Writing Image to Excel:\n{image_path}")
    img = Image(image_path)
    column_letter = get_column_letter(pos[1] + 1)
    row = pos[0] + 1
    ws.add_image(img, f'{column_letter}{row}')


def run_macro(macro_name, arguments):
    temp_file_path = Path(tempfile.gettempdir()) / 'workbookargs.json'
    with open(temp_file_path, 'w') as f:
        f.write(json.dumps(arguments))

    xl = win32com.client.Dispatch("Excel.Application")
    xl.Workbooks.Open(Filename=Path(Path(__file__).parent / 'templates/macros.xlsm'))
    xl.Application.Run(macro_name)
    xl.Application.Quit()

class MacroRunner:
    def __init__(self, macros_path=None, visible=False):
        self.macros_path = Path(macros_path) if macros_path else Path(__file__).parent / 'templates/macros.xlsm'
        self.visible = visible
        self.xl = None
        self.macro_wb = None
        self.args_path = Path(tempfile.gettempdir()) / 'workbookargs.json'
        self._start()

    def __enter__(self):
        if self.xl is None:
            self._start()
        return self

    def __exit__(self, exc_type, exc, tb):
        self.close()

    def _start(self):
        # Isolated Excel instance
        self.xl = win32com.client.DispatchEx("Excel.Application")
        self.xl.Visible = self.visible
        self.xl.ScreenUpdating = False
        self.xl.DisplayAlerts = False
        self.xl.EnableEvents = False
        try:
            self.xl.AutoRecover.Enabled = False
        except Exception:
            pass
        try:
            self.xl.AskToUpdateLinks = False
        except Exception:
            pass

        # Open the macros workbook once and reuse
        self.macro_wb = self.xl.Workbooks.Open(
            Filename=str(self.macros_path),
            ReadOnly=True,
            UpdateLinks=0,
            AddToMru=False,
        )

    def _write_args(self, arguments):
        # Atomic write so VBA never reads a partial file
        fd, tmp_name = tempfile.mkstemp(prefix="workbookargs_", suffix=".json")
        tmp_path = Path(tmp_name)
        try:
            with open(fd, "w", encoding="utf-8") as f:
                f.write(json.dumps(arguments))
                f.flush()
            # Replace into place atomically
            tmp_path.replace(self.args_path)
        finally:
            # If replace failed, ensure tmp is cleaned up
            if tmp_path.exists() and tmp_path != self.args_path:
                try:
                    tmp_path.unlink()
                except Exception:
                    pass

    def run(self, macro_name, arguments=None, cleanup_args=True):
        if arguments is None:
            arguments = {}
        self._write_args(arguments)

        # Redirect temp artifacts away from the report folder
        temp_dir = Path(tempfile.gettempdir())
        prev_cwd = Path.cwd()
        try:
            # 1) Change Python process CWD (some Excel/VBA calls use the process cwd)
            os.chdir(temp_dir)

            # 2) Tell Excel to use temp dir for its internal file operations
            try:
                # ChangeFileOpenDirectory affects many internal operations
                self.xl.ChangeFileOpenDirectory(str(temp_dir))
            except Exception:
                pass
            try:
                # DefaultFilePath is used by some dialogs/background operations
                self.xl.DefaultFilePath = str(temp_dir)
            except Exception:
                pass

            # Run synchronously
            self.xl.Run(macro_name)
        finally:
            # Restore cwd and Excel directory settings
            try:
                os.chdir(prev_cwd)
            except Exception:
                pass
            try:
                self.xl.ChangeFileOpenDirectory(str(prev_cwd))
            except Exception:
                pass

        if cleanup_args:
            try:
                self.args_path.unlink(missing_ok=True)
            except TypeError:
                if self.args_path.exists():
                    self.args_path.unlink()

    def close(self):
        # Close the macros workbook and Excel cleanly
        if self.macro_wb is not None:
            try:
                self.macro_wb.Close(SaveChanges=False)
            except Exception:
                pass
            self.macro_wb = None
        if self.xl is not None:
            try:
                self.xl.Quit()
            except Exception:
                pass
            self.xl = None
        # Encourage COM cleanup
        try:
            import gc
            gc.collect()
        except Exception:
            pass

class ReportCreator:
    def __init__(self, filename=None, request_path=None, template=None, *args, **kwargs):
        if filename is None:
            filename = Path('reports') / f'{Path.cwd().name}_{datetime.today().strftime('%d%b%Y').upper()}.xlsx'
        self.filename = Path(filename)
        self.request_path = None
        if request_path is not None:
            self.request_path = Path(request_path)
        self.wb = self.load_workbook(filename, template=template, *args, **kwargs)
        self.ws = self.wb.active

    def create_sheet(self, name, title=None):
        sheet_name = name[:30]
        self.wb.create_sheet(sheet_name)
        self.ws = self.wb[sheet_name]
        self.wb.active = self.ws
        if title:
            self.set_page_title(title)
        else:
            self.set_page_title(name)
        return self.ws

    def load_workbook(self, filename, template=None, *args, **kwargs):
        if template is None:
            template = Path(__file__).parent / 'templates/template.xlsx'
        else:
            template = Path(template)
        destination = Path(filename)
        destination.write_bytes(template.read_bytes())
        return openpyxl.load_workbook(filename, *args, **kwargs)

    def set_page_title(self, title, pos=None):
        logging.info(f"Writing Title to Excel:\n{title}")
        if pos is None:
            pos = [0, 0]
        cell = self.ws.cell(row=pos[0] + 1, column=pos[1] + 1, value=title)
        cell.style = 'SAS page title'

    def add_note(self, note, position=None, merge=None):
        logging.info(f"Writing Note to Excel:\n{note}")
        if position is None:
            position = [1, 1]
        position[0] += 1
        position[1] += 1
        cell = self.ws.cell(row=position[0], column=position[1], value=note)
        if merge:
            self.ws.merge_cells(start_row=position[0],
                                start_column=position[1],
                                end_row=position[0] + merge[0],
                                end_column=position[1] + merge[1])
        cell.style = 'SAS note'

    def add_title(self, title, position=None):
        logging.info(f"Writing Title to Excel:\n{title}")
        if position is None:
            position = [1, 1]
        position[0] += 1
        position[1] += 1
        cell = self.ws.cell(row=position[0], column=position[1], value=title)
        cell.style = 'SAS page title'

    def get_contiguous_sequences(self, codes):
        previous_index = [0 for _ in codes[0]]
        final_position = len(codes[0]) - 1
        groups = []
        for col in codes:
            previous_code = col[0]
            start = 0
            sub_groups = []
            for pos, code in enumerate(col):
                if pos != 0:
                    if previous_code != code or previous_index[pos] != previous_index[pos - 1]:
                        if pos - start > 1:
                            sub_groups.append([start, pos - 1])
                        start = pos
                previous_code = code
            if final_position - start > 0:
                sub_groups.append([start, final_position])
            groups.append(sub_groups)
            previous_index = col
        return groups

    def get_index_codes(self, index):
        if index.nlevels > 1:
            return index.codes
        else:
            unique_values = index.unique().to_list()
            return [[unique_values.index(value) for value in index]]

    def write_table(self, df, table_start_pos=None, write_index=True, write_headers=True, title=None, percent=False):
        logging.info(f"Writing DataFrame to Excel:\n{df.to_string()}")
        df = df.copy()
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = pd.MultiIndex.from_tuples(
                tuple(str(item) for item in level) for level in df.columns
            )
        else:
            df.columns = df.columns.astype(str)

        if table_start_pos is None:
            table_start_pos = [0, 0]

        table_original_start = table_start_pos

        if write_index:
            index_levels = df.index.nlevels
            index_groupings = get_contiguous_sequences(get_index_codes(df.index))
        else:
            index_levels = 0
            index_groupings = []

        if write_headers:
            column_levels = df.columns.nlevels
            header_groupings = get_contiguous_sequences(get_index_codes(df.columns))
        else:
            column_levels = 0
            header_groupings = []

        if title:
            cell = self.ws.cell(row=table_start_pos[0] + 1, column=table_start_pos[1] + 1, value=title)
            cell.style = 'SAS Title'
            cell.alignment = Alignment(horizontal='center', vertical='center')
            self.ws.merge_cells(start_row=table_start_pos[0] + 1,
                                start_column=table_start_pos[1] + 1,
                                end_row=table_start_pos[0] + 1,
                                end_column=table_start_pos[1] + index_levels + df.shape[1])
            table_start_pos[0] += 1

        if write_index:
            df_temp = df.reset_index()
        else:
            df_temp = df

        for column_n, (headers, column) in enumerate(df_temp.to_dict().items()):
            if isinstance(headers, str):
                headers = [headers]
            is_index = column_n < index_levels

            current_column = column_n + table_start_pos[1] + 1
            for row_n, value in enumerate(column.values()):
                if row_n == 0 and write_headers:
                    for header_pos, header in enumerate(headers):
                        if is_index and (headers[0][:5] in ['level', 'index']):
                            header = ''
                        cell = self.ws.cell(row=table_start_pos[0] + 1 + header_pos, column=current_column,
                                            value=header)
                        cell.style = 'SAS Title'
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                current_row = row_n + table_start_pos[0] + 1 + column_levels
                cell = self.ws.cell(row=current_row, column=current_column, value=value)
                if is_index:
                    cell.style = 'SAS Title'
                else:
                    cell.style = 'SAS Data'
                    if percent:
                        cell.number_format = numbers.FORMAT_PERCENTAGE_00

        for i, index_grouping in enumerate(index_groupings):
            for group in index_grouping:
                self.ws.merge_cells(start_row=table_start_pos[0] + 1 + group[0] + column_levels,
                                    start_column=table_start_pos[1] + 1 + i,
                                    end_row=table_start_pos[0] + 1 + group[1] + column_levels,
                                    end_column=table_start_pos[1] + 1 + i)

        for i, index_grouping in enumerate(header_groupings):
            for group in index_grouping:
                self.ws.merge_cells(start_row=table_start_pos[0] + 1,
                                    start_column=table_start_pos[1] + 1 + i + group[0] + index_levels,
                                    end_row=table_start_pos[0] + 1,
                                    end_column=table_start_pos[1] + 1 + i + group[1] + index_levels)

        if column_levels > 1 or index_levels > 1:
            # Untested. But should work.
            self.ws.merge_cells(start_row=table_start_pos[0] + 1,
                                start_column=table_start_pos[1] + 1,
                                end_row=table_start_pos[0] + column_levels,
                                end_column=table_start_pos[1] + index_levels)

        # Makes cells background color change

        # for column in self.ws.columns:
        #     max_length = 0
        #     column = [cell for cell in column if cell.value]  # Checking if the cell has a value
        #     for cell in column:
        #         try:  # Necessary to avoid error on empty cells
        #             if len(str(cell.value)) > max_length:
        #                 max_length = len(cell.value)
        #         except:
        #             pass
        #     adjusted_width = (max_length + 2)  # Additional +2 for a bit of margin
        #     try:
        #         self.ws.column_dimensions[column[0].column_letter].width = min(adjusted_width, 100)
        #     except:
        #         pass

        return [table_original_start,
                [table_start_pos[0] + column_levels + (1 if title else 0) + df.shape[0],
                 table_start_pos[1] + index_levels + df.shape[1]]]

    def insert_image(self, image_path, pos):
        logging.info(f"Writing Image to Excel:\n{image_path}")
        img = Image(image_path)
        column_letter = get_column_letter(pos[1] + 1)
        row = pos[0] + 1
        self.ws.add_image(img, f'{column_letter}{row}')

    def save(self):
        self.wb.active = 0
        self.wb.save(self.filename)
        mr = MacroRunner()
        mr.run('AutoFitAllColumns', {'workbook_path': str(self.filename.absolute())})
        if self.request_path:
            mr.run('EmbedWordDocument', {"workbook_path": str(self.filename.absolute()),
                                         "embed_path": str(self.request_path.absolute()),
                                         "target_sheet": "Request",
                                         "target_cell": "A1"})
        mr.close()
        print(f"Saving to {str(self.filename.absolute())}")
